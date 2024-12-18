import logging
import math
import os
from typing import Dict

import pandas as pd
from dotenv import load_dotenv
from fastapi import APIRouter, Request
from fastapi.responses import RedirectResponse
from fastapi.security import OAuth2PasswordBearer
from pydantic import Field, BaseModel

from src.app.user.schema import OTPVerificationRequest, loginwithemailandpassword, TOTPVerificationRequest, NumbersInput
from src.app.user.service import initiate_oauth, handle_oauth_callback, create_and_send_otp, verify_otp, \
    login_with_totp_service, verify_totp_service, read_and_compute_sum
from src.connection.connection import ConnectionHandler, get_connection_handler_for_app
from fastapi import APIRouter, HTTPException, Depends


oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# Initialize logger
logger = logging.getLogger(__name__)

load_dotenv()
user_router = APIRouter()


@user_router.get("/login")
async def login(request: Request):
    """Redirect user to Google OAuth login page"""
    try:
        authorization_url = await initiate_oauth(request)

        return RedirectResponse(url=authorization_url)

    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"OAuth initialization failed: {str(e)}")


@user_router.get("/auth/callback")
async def auth_callback(request: Request):
    """Handle OAuth callback from Google"""
    try:
        user = await handle_oauth_callback(request=request)
        return {
            "message": "Login successful",
            "user": user
        }

    except Exception as e:
        logger.error(f"Callback error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Authentication failed: {str(e)}")


@user_router.post("/login_otp")
async def login_with_otp(
        mobile_number: str,
        connection_handler: ConnectionHandler = Depends(get_connection_handler_for_app),
):
    session = connection_handler.session

    # Await the async function and get the result
    result = await create_and_send_otp(mobile_number, session)

    if isinstance(result, dict):
        return result
    else:
        return {"message": "Internal server error", "status": "error"}


@user_router.post("/verify_otp")
async def login_with_otp(
        request: OTPVerificationRequest,
        connection_handler: ConnectionHandler = Depends(get_connection_handler_for_app),
):
    session = connection_handler.session
    tokens = await verify_otp(request.mobile_number, request.otp, session)
    response={"message": "OTP verified successfully", **tokens}
    return response
@user_router.post("/login_totp")
async def login_with_totp(
    request:loginwithemailandpassword,
    connection_handler: ConnectionHandler = Depends(get_connection_handler_for_app),
):
    session = connection_handler.session
    result = await login_with_totp_service(request.mobile_number, request.password, session)
    return result

@user_router.post("/verify_totp")
async def verify_totp(
    request: TOTPVerificationRequest,
    connection_handler: ConnectionHandler = Depends(get_connection_handler_for_app),
):
    session = connection_handler.session

    result = await verify_totp_service(request.mobile_number, request.totp, session)

    return result


@user_router.post("/retrieve/sum")
async def excel_sum(payload: Dict[str, float], csv_path: str = "/home/mind/Downloads/test_sum.csv"):
    try:
        no1 = payload.get("no1")
        no2 = payload.get("no2")

        if no1 is None or no2 is None:
            raise HTTPException(status_code=400, detail="Missing no1 or no2 in the payload")

        sum_result = read_and_compute_sum(csv_path, no1, no2)

        return {"sum": sum_result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing the request: {str(e)}")


EXCEL_FILE_PATH = '/home/mind/Documents/final_test.xlsx'

def ensure_csv_exists():
    if not os.path.exists(EXCEL_FILE_PATH):
        df = pd.DataFrame(columns=['no1', 'no2', 'sum'])
        df.to_csv(EXCEL_FILE_PATH, index=False)


from openpyxl import load_workbook
from fastapi import HTTPException

class NumbersInput(BaseModel):
    no1: float
    no2: float


@user_router.post("/retrieve/sum/2")
async def excel_sum(input_data: NumbersInput):
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # Find next empty row
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                break
        else:
            row = ws.max_row + 1

        ws.cell(row=row, column=1, value=input_data.no1)
        ws.cell(row=row, column=2, value=input_data.no2)

        wb.save(EXCEL_FILE_PATH)
        wb.close()

        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        ws = wb.active

        sum_value = ws.cell(row=row, column=3).value

        return {
            "no1": input_data.no1,
            "no2": input_data.no2,
            "sum": sum_value
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@user_router.get("/retrieve/sum/3")
async def get_excel_data():
    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)  # Use `data_only=True` to get calculated formula values
        ws = wb.active

        no1 = ws["A2"].value
        no2 = ws["B2"].value
        sum_value = ws["C2"].value

        wb.close()

        return {
            "no1": no1,
            "no2": no2,
            "sum": sum_value
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))